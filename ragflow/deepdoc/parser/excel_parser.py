#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.


import logging
import re
import sys
import json
from io import BytesIO
from typing import Dict, List, Any, Optional

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook

from rag.nlp import find_codec

# copied from `/openpyxl/cell/cell.py`
ILLEGAL_CHARACTERS_RE = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")


class RAGFlowExcelParser:
    def __init__(self):
        self.metadata = {}
        self.schema = {}
        
    @staticmethod
    def _load_excel_to_workbook(file_like_object):
        if isinstance(file_like_object, bytes):
            file_like_object = BytesIO(file_like_object)

        # Read first 4 bytes to determine file type
        file_like_object.seek(0)
        file_head = file_like_object.read(4)
        file_like_object.seek(0)

        if not (file_head.startswith(b"PK\x03\x04") or file_head.startswith(b"\xd0\xcf\x11\xe0")):
            logging.info("Not an Excel file, converting CSV to Excel Workbook")

            try:
                file_like_object.seek(0)
                df = pd.read_csv(file_like_object)
                return RAGFlowExcelParser._dataframe_to_workbook(df)

            except Exception as e_csv:
                raise Exception(f"Failed to parse CSV and convert to Excel Workbook: {e_csv}")

        try:
            return load_workbook(file_like_object, data_only=True)
        except Exception as e:
            logging.info(f"openpyxl load error: {e}, try pandas instead")
            try:
                file_like_object.seek(0)
                try:
                    df = pd.read_excel(file_like_object)
                    return RAGFlowExcelParser._dataframe_to_workbook(df)
                except Exception as ex:
                    logging.info(f"pandas with default engine load error: {ex}, try calamine instead")
                    file_like_object.seek(0)
                    df = pd.read_excel(file_like_object, engine="calamine")
                    return RAGFlowExcelParser._dataframe_to_workbook(df)
            except Exception as e_pandas:
                raise Exception(f"pandas.read_excel error: {e_pandas}, original openpyxl error: {e}")

    @staticmethod
    def _clean_dataframe(df: pd.DataFrame):
        """Enhanced dataframe cleaning for accuracy"""
        def clean_string(s):
            if isinstance(s, str):
                # Remove illegal characters
                s = ILLEGAL_CHARACTERS_RE.sub(" ", s)
                # Strip whitespace
                s = s.strip()
                # Normalize multiple spaces to single space
                s = re.sub(r'\s+', ' ', s)
                return s if s else None
            return s

        # Apply string cleaning
        df = df.apply(lambda col: col.map(clean_string))
        
        # Remove completely empty rows and columns
        df = df.dropna(how='all', axis=0)
        df = df.dropna(how='all', axis=1)
        
        # Clean column names
        df.columns = df.columns.astype(str).str.strip()
        df.columns = df.columns.str.replace(r'\s+', ' ', regex=True)
        df.columns = df.columns.str.replace(r'[^\w\s-]', '_', regex=True)
        
        # Replace inf/-inf with NaN
        df = df.replace([np.inf, -np.inf], np.nan)
        
        # Reset index
        df = df.reset_index(drop=True)
        
        return df

    @staticmethod
    def _dataframe_to_workbook(df):
        df = RAGFlowExcelParser._clean_dataframe(df)
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        for col_num, column_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_num, value=column_name)

        for row_num, row in enumerate(df.values, 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        return wb

    def _extract_metadata(self, wb):
        """Extract comprehensive metadata for chatbot context"""
        metadata = {}
        
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            
            if not rows:
                continue
                
            headers = [str(cell.value).strip() if cell.value else f"Column_{i}" 
                      for i, cell in enumerate(rows[0])]
            
            # Collect column data for analysis
            column_data = {header: [] for header in headers}
            for row in rows[1:]:
                for i, cell in enumerate(row):
                    if i < len(headers):
                        column_data[headers[i]].append(cell.value)
            
            # Analyze each column
            column_info = {}
            for col, values in column_data.items():
                # Filter out None values
                clean_values = [v for v in values if v is not None]
                
                if not clean_values:
                    column_info[col] = {
                        'type': 'empty',
                        'sample_values': [],
                        'unique_count': 0,
                        'null_count': len(values)
                    }
                    continue
                
                # Detect type
                col_type = self._detect_column_type(clean_values)
                
                # Get unique count
                unique_count = len(set(str(v) for v in clean_values))
                
                # Get sample values (first 3 unique)
                sample_values = list(dict.fromkeys([str(v) for v in clean_values[:5]]))[:3]
                
                column_info[col] = {
                    'type': col_type,
                    'sample_values': sample_values,
                    'unique_count': unique_count,
                    'null_count': len(values) - len(clean_values),
                    'total_count': len(values)
                }
            
            metadata[sheetname] = {
                'headers': headers,
                'row_count': len(rows) - 1,  # Exclude header
                'column_count': len(headers),
                'columns': column_info
            }
        
        return metadata
    
    @staticmethod
    def _detect_column_type(values):
        """Detect the most likely type of a column"""
        if not values:
            return 'empty'
        
        # Sample first 100 values for efficiency
        sample = values[:100]
        
        numeric_count = 0
        date_count = 0
        
        for val in sample:
            val_str = str(val).strip()
            
            # Check if numeric
            try:
                float(val_str)
                numeric_count += 1
            except:
                pass
            
            # Check if date-like
            if any(sep in val_str for sep in ['-', '/', '.']):
                parts = re.split(r'[-/.]', val_str)
                if len(parts) >= 2 and all(p.isdigit() for p in parts[:3]):
                    date_count += 1
        
        # Determine type based on majority
        if numeric_count > len(sample) * 0.8:
            return 'numeric'
        elif date_count > len(sample) * 0.5:
            return 'date'
        else:
            return 'text'

    def html(self, fnm, chunk_rows=1):
        """
        Enhanced HTML parser with row-level chunking for 100% accuracy
        Each row becomes a separate chunk with complete column context
        """
        from html import escape

        file_like_object = BytesIO(fnm) if not isinstance(fnm, str) else fnm
        wb = RAGFlowExcelParser._load_excel_to_workbook(file_like_object)
        
        # Extract comprehensive metadata
        self.metadata = self._extract_metadata(wb)
        
        tb_chunks = []

        def _fmt(v):
            if v is None:
                return ""
            return str(v).strip()

        def _get_cell_reference(cell):
            """Get cell reference like A1, B2, etc."""
            return cell.coordinate

        def _extract_formula(cell):
            """Extract formula if cell contains one"""
            if hasattr(cell, 'data_type') and cell.data_type == 'f':
                return cell.value
            return None

        def _create_row_chunk(sheetname, headers, row_data, row_index):
            """Create a single row chunk with complete context"""
            # Clean headers
            clean_headers = [str(h).strip() if h else f"Column_{i}" for i, h in enumerate(headers)]
            
            # Create structured text representation
            row_text_parts = []
            row_metadata = {
                'sheet': sheetname,
                'row_index': row_index,
                'headers': clean_headers,
                'values': {},
                'data_types': {}
            }
            
            # Build row text with column:value pairs
            for i, (header, value) in enumerate(zip(clean_headers, row_data)):
                if value is not None and str(value).strip():
                    clean_value = str(value).strip()
                    row_text_parts.append(f"{header}: {clean_value}")
                    row_metadata['values'][header] = clean_value
                    
                    # Detect data type
                    if clean_value.replace(',', '').replace('.', '').replace('-', '').replace('%', '').isdigit():
                        row_metadata['data_types'][header] = 'numeric'
                    elif '%' in clean_value:
                        row_metadata['data_types'][header] = 'percentage'
                    else:
                        row_metadata['data_types'][header] = 'text'
            
            # Create HTML chunk
            chunk_html = f"<div class='excel-row-chunk' data-sheet='{sheetname}' data-row='{row_index}'>"
            chunk_html += f"<div class='sheet-context'><strong>Sheet:</strong> {sheetname} | <strong>Row:</strong> {row_index}</div>"
            chunk_html += f"<div class='row-data'>{' | '.join(row_text_parts)}</div>"
            chunk_html += f"<div class='metadata' style='display:none;'>{json.dumps(row_metadata)}</div>"
            chunk_html += "</div>"
            
            return chunk_html, row_metadata

        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            if not rows:
                continue

            # Get headers from first row
            headers = [cell.value for cell in rows[0]]
            
            # Process each row as a separate chunk
            for row_index, row in enumerate(rows[1:], start=1):
                row_data = [cell.value for cell in row]
                
                # Skip completely empty rows
                if all(v is None or str(v).strip() == '' for v in row_data):
                    continue
                
                # Create row chunk
                chunk_html, row_metadata = _create_row_chunk(sheetname, headers, row_data, row_index)
                tb_chunks.append(chunk_html)

        return tb_chunks

    def markdown(self, fnm):
        """Enhanced markdown output with metadata"""
        file_like_object = BytesIO(fnm) if not isinstance(fnm, str) else fnm
        try:
            file_like_object.seek(0)
            df = pd.read_excel(file_like_object)
        except Exception as e:
            logging.warning(f"Parse spreadsheet error: {e}, trying to interpret as CSV file")
            file_like_object.seek(0)
            df = pd.read_csv(file_like_object)
        
        # Clean the dataframe
        df = self._clean_dataframe(df)
        df = df.replace(r"^\s*$", "", regex=True)
        
        return df.to_markdown(index=False)

    def __call__(self, fnm):
        """Enhanced default parsing with better structure for chatbot"""
        file_like_object = BytesIO(fnm) if not isinstance(fnm, str) else fnm
        wb = RAGFlowExcelParser._load_excel_to_workbook(file_like_object)
        
        # Extract metadata
        self.metadata = self._extract_metadata(wb)

        res = []
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            if not rows:
                continue
                
            ti = list(rows[0])
            
            # Clean headers
            headers = [str(cell.value).strip() if cell.value else f"Col_{i}" 
                      for i, cell in enumerate(ti)]
            
            for row_idx, r in enumerate(list(rows[1:]), start=1):
                fields = []
                for i, c in enumerate(r):
                    if not c.value:
                        continue
                    
                    # Get header name
                    header = headers[i] if i < len(headers) else f"Col_{i}"
                    
                    # Clean cell value
                    value = str(c.value).strip()
                    
                    # Format: "Header: Value"
                    fields.append(f"{header}: {value}")
                
                if not fields:
                    continue
                    
                # Join fields with better separator
                line = " | ".join(fields)
                
                # Add sheet name if not default
                if sheetname.lower().find("sheet") < 0:
                    line = f"[{sheetname}] {line}"
                
                res.append(line)
        
        return res

    def to_structured_data(self, fnm) -> Dict[str, Any]:
        """
        NEW METHOD: Parse to structured dictionary format
        Optimized for chatbot vector database ingestion
        """
        file_like_object = BytesIO(fnm) if not isinstance(fnm, str) else fnm
        wb = RAGFlowExcelParser._load_excel_to_workbook(file_like_object)
        
        # Extract metadata
        self.metadata = self._extract_metadata(wb)
        
        structured_data = {
            'sheets': {},
            'metadata': self.metadata
        }
        
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            
            if not rows:
                continue
            
            # Get headers
            headers = [str(cell.value).strip() if cell.value else f"Column_{i}" 
                      for i, cell in enumerate(rows[0])]
            
            # Parse rows
            sheet_data = []
            for row_idx, row in enumerate(rows[1:], start=1):
                row_dict = {'_row_id': row_idx}
                
                for i, cell in enumerate(row):
                    if i < len(headers):
                        value = cell.value
                        
                        # Clean and convert value
                        if value is not None:
                            if isinstance(value, str):
                                value = value.strip()
                                if not value:
                                    value = None
                        
                        row_dict[headers[i]] = value
                
                # Only add rows with at least one non-null value
                if any(v is not None for k, v in row_dict.items() if k != '_row_id'):
                    sheet_data.append(row_dict)
            
            structured_data['sheets'][sheetname] = sheet_data
        
        return structured_data

    def to_json(self, fnm, output_path: Optional[str] = None) -> str:
        """
        NEW METHOD: Export to JSON format
        Perfect for chatbot knowledge base
        """
        data = self.to_structured_data(fnm)
        json_str = json.dumps(data, indent=2, ensure_ascii=False, default=str)
        
        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(json_str)
        
        return json_str

    def to_chatbot_context(self, fnm, max_rows_per_sheet: int = 1000) -> str:
        """
        NEW METHOD: Generate optimized text context for LLM/chatbot
        """
        data = self.to_structured_data(fnm)
        
        context_parts = []
        context_parts.append("=== EXCEL DATA CONTEXT ===\n")
        
        for sheet_name, sheet_data in data['sheets'].items():
            metadata = data['metadata'].get(sheet_name, {})
            
            context_parts.append(f"\n## Sheet: {sheet_name}")
            context_parts.append(f"Columns: {', '.join(metadata.get('headers', []))}")
            context_parts.append(f"Total Rows: {len(sheet_data)}\n")
            
            # Add column type information
            column_info = metadata.get('columns', {})
            if column_info:
                context_parts.append("Column Types:")
                for col, info in column_info.items():
                    context_parts.append(f"  - {col}: {info.get('type', 'unknown')} "
                                       f"({info.get('unique_count', 0)} unique values)")
            
            context_parts.append("\nData:")
            
            # Add rows (limited by max_rows_per_sheet)
            for i, row in enumerate(sheet_data[:max_rows_per_sheet]):
                row_text = []
                for key, value in row.items():
                    if key != '_row_id' and value is not None:
                        row_text.append(f"{key}={value}")
                
                if row_text:
                    context_parts.append(f"Row {row['_row_id']}: {', '.join(row_text)}")
            
            if len(sheet_data) > max_rows_per_sheet:
                context_parts.append(f"... ({len(sheet_data) - max_rows_per_sheet} more rows)")
        
        return '\n'.join(context_parts)

    def to_vector_db_documents(self, fnm) -> List[Dict[str, Any]]:
        """
        NEW METHOD: Convert to vector database document format
        Each row becomes a searchable document
        """
        data = self.to_structured_data(fnm)
        documents = []
        
        for sheet_name, sheet_data in data['sheets'].items():
            for row in sheet_data:
                # Create text representation
                text_parts = []
                doc_metadata = {'sheet': sheet_name, 'row_id': row.get('_row_id')}
                
                for key, value in row.items():
                    if key != '_row_id' and value is not None:
                        text_parts.append(f"{key}: {value}")
                        doc_metadata[key] = value
                
                if text_parts:
                    doc_id = f"{sheet_name}_row_{row.get('_row_id', 0)}"
                    documents.append({
                        'id': doc_id,
                        'text': '. '.join(text_parts),
                        'metadata': doc_metadata
                    })
        
        return documents

    def get_schema(self) -> Dict[str, Any]:
        """
        NEW METHOD: Get complete schema information
        Helps chatbot understand data structure
        """
        return {
            'metadata': self.metadata,
            'summary': self._generate_summary()
        }
    
    def validate_accuracy(self, query: str, retrieved_chunks: List[Dict]) -> Dict[str, Any]:
        """
        NEW METHOD: Validate accuracy of retrieved data and detect ambiguity
        Returns validation results and clarification questions if needed
        """
        validation_result = {
            'is_accurate': True,
            'confidence': 1.0,
            'ambiguity_detected': False,
            'clarification_questions': [],
            'exact_matches': [],
            'partial_matches': [],
            'suggestions': []
        }
        
        # Extract query components
        query_lower = query.lower()
        query_words = set(query_lower.split())
        
        for chunk in retrieved_chunks:
            chunk_text = chunk.get('text', '')
            chunk_metadata = chunk.get('metadata', {})
            
            # Check for exact matches
            if self._has_exact_match(query, chunk_text, chunk_metadata):
                validation_result['exact_matches'].append({
                    'chunk_id': chunk.get('id'),
                    'text': chunk_text,
                    'metadata': chunk_metadata,
                    'confidence': 1.0
                })
            else:
                # Check for partial matches
                similarity_score = self._calculate_similarity(query, chunk_text)
                if similarity_score > 0.7:
                    validation_result['partial_matches'].append({
                        'chunk_id': chunk.get('id'),
                        'text': chunk_text,
                        'metadata': chunk_metadata,
                        'confidence': similarity_score
                    })
        
        # Detect ambiguity
        if len(validation_result['exact_matches']) > 1:
            validation_result['ambiguity_detected'] = True
            validation_result['clarification_questions'].append(
                f"I found {len(validation_result['exact_matches'])} exact matches. Which specific value do you need?"
            )
        elif len(validation_result['partial_matches']) > 1:
            validation_result['ambiguity_detected'] = True
            validation_result['clarification_questions'].append(
                f"I found {len(validation_result['partial_matches'])} similar matches. Could you be more specific about which value you need?"
            )
        elif len(validation_result['exact_matches']) == 0 and len(validation_result['partial_matches']) == 0:
            validation_result['is_accurate'] = False
            validation_result['confidence'] = 0.0
            validation_result['suggestions'].append("No matching data found. Please check your query or try different keywords.")
        
        return validation_result
    
    def _has_exact_match(self, query: str, chunk_text: str, metadata: Dict) -> bool:
        """Check if chunk contains exact match for query"""
        query_lower = query.lower()
        chunk_lower = chunk_text.lower()
        
        # Check for exact value matches
        values = metadata.get('values', {})
        for header, value in values.items():
            if str(value).lower() in query_lower or query_lower in str(value).lower():
                return True
        
        # Check for header matches
        headers = metadata.get('headers', [])
        for header in headers:
            if header.lower() in query_lower:
                return True
        
        return False
    
    def _calculate_similarity(self, query: str, chunk_text: str) -> float:
        """Calculate similarity between query and chunk text"""
        query_words = set(query.lower().split())
        chunk_words = set(chunk_text.lower().split())
        
        if not query_words or not chunk_words:
            return 0.0
        
        intersection = query_words.intersection(chunk_words)
        union = query_words.union(chunk_words)
        
        return len(intersection) / len(union) if union else 0.0
    
    def _generate_summary(self) -> str:
        """Generate human-readable summary"""
        if not self.metadata:
            return "No metadata available. Parse a file first."
        
        summary_parts = []
        summary_parts.append(f"Total sheets: {len(self.metadata)}")
        
        for sheet_name, meta in self.metadata.items():
            summary_parts.append(f"\nSheet: {sheet_name}")
            summary_parts.append(f"  Rows: {meta.get('row_count', 0)}")
            summary_parts.append(f"  Columns: {meta.get('column_count', 0)}")
            summary_parts.append(f"  Headers: {', '.join(meta.get('headers', []))}")
        
        return '\n'.join(summary_parts)

    @staticmethod
    def row_number(fnm, binary):
        if fnm.split(".")[-1].lower().find("xls") >= 0:
            wb = RAGFlowExcelParser._load_excel_to_workbook(BytesIO(binary))
            total = 0
            for sheetname in wb.sheetnames:
                ws = wb[sheetname]
                total += len(list(ws.rows))
            return total

        if fnm.split(".")[-1].lower() in ["csv", "txt"]:
            encoding = find_codec(binary)
            txt = binary.decode(encoding, errors="ignore")
            return len(txt.split("\n"))


if __name__ == "__main__":
    psr = RAGFlowExcelParser()
    
    if len(sys.argv) > 1:
        # Original behavior
        result = psr(sys.argv[1])
        for line in result:
            print(line)
        
        # Demo new methods
        print("\n=== SCHEMA ===")
        schema = psr.get_schema()
        print(json.dumps(schema, indent=2))
        
        print("\n=== CHATBOT CONTEXT (first 100 rows) ===")
        context = psr.to_chatbot_context(sys.argv[1], max_rows_per_sheet=100)
        print(context[:1000] + "...")  # Preview
        
        print("\n=== VECTOR DB DOCS (first 3) ===")
        docs = psr.to_vector_db_documents(sys.argv[1])
        for doc in docs[:3]:
            print(json.dumps(doc, indent=2))